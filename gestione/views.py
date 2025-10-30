# gestione/views.py

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.db import models 
from django.db.models import Sum, Count, F, ExpressionWrapper, DecimalField, Q, Subquery, OuterRef, Avg, Case, When, Value
from django.db.models.functions import TruncMonth, Coalesce
from decimal import Decimal
from .models import (
    Vendita, CategoriaMerceologica, CategoriaServizio, StatMensile, Budget, 
    Trattativa, Attivita, MessaggioChat, ProfiloUtente,
    ImpostazioniGenerali, RuoloCosto # Assicurati che RuoloCosto sia importato
)
import datetime
from django.contrib.auth.models import User
from django.utils import timezone
import json
from .forms import (
    TrattativaVintaForm, AttivitaForm, MessaggioChatForm, 
    TrattativaForm, StatMensileForm, TrattativaDettaglioForm
)
from django.contrib import messages
from django.views.decorators.http import require_POST
from django.http import HttpResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# --- SOGLIE DI ALLERTA ---
ALERT_SOGLIA_GIORNI_EVASIONE = 60
ALERT_SOGLIA_RESI_DIFETTI = 5
ALERT_SOGLIA_FINAN_NON_APPROV = 3

# --- FUNZIONI HELPER (BASATE SU RUOLO) ---
def get_costo_attivita_query():
    """
    Ritorna un'espressione SQL per calcolare il costo delle attività.
    Logica: Ore * costo_orario del Ruolo.
    """
    return Sum(
        F('tempo_dedicato_ore') * Coalesce(F('ruolo__costo_orario'), Decimal(0.0)),
        output_field=DecimalField()
    )

def get_costo_personale_trattativa(trattativa_id):
    """Calcola il costo totale del personale per una singola trattativa."""
    costo = Attivita.objects.filter(
        trattativa_id=trattativa_id
    ).aggregate(
        costo_totale=Coalesce(get_costo_attivita_query(), Decimal(0.0))
    )['costo_totale']
    return costo

def get_costo_personale_query():
    """
    Ritorna un'espressione SQL per calcolare il costo aggregato.
    """
    return Sum(
        F('attivita__tempo_dedicato_ore') * Coalesce(F('attivita__ruolo__costo_orario'), Decimal(0.0)),
        output_field=DecimalField()
    )
    
# --- FINE FUNZIONI HELPER ---


# --- NUOVA FUNZIONE HELPER (PUNTO 19) ---
def get_trattative_annotate():
    tutte_le_trattative = Trattativa.objects.all().select_related('commerciale')

    # 1. Annotiamo il costo del personale (da Attività)
    trattative_con_costi = tutte_le_trattative.annotate(
        costo_personale_totale=Coalesce(
            Sum(
                F('attivita__tempo_dedicato_ore') * Coalesce(F('attivita__ruolo__costo_orario'), Decimal(0.0)),
                output_field=DecimalField()
            ),
            Decimal(0.0) 
        ),
        ricavo_servizi_totale=Coalesce(
            Sum('attivita__prezzo_vendita_attivita'),
            Decimal(0.0)
        )
    )

    # 3. Annotiamo il margine e la percentuale
    trattative_con_margine = trattative_con_costi.annotate(
        valore_totale_stimato=F('valore_stimato') + F('ricavo_servizi_totale'),
        costo_totale_stimato=F('costo_materiali_stimato') + F('costo_personale_totale'),
        margine_stimato_euro=F('valore_totale_stimato') - F('costo_totale_stimato'),
        margine_stimato_perc=Case(
            When(valore_totale_stimato=0, then=Value(Decimal(0.0))),
            default=ExpressionWrapper(
                ( (F('valore_totale_stimato') - F('costo_totale_stimato')) * Decimal(100.0) ) / F('valore_totale_stimato'),
                output_field=DecimalField(max_digits=5, decimal_places=2)
            ),
            output_field=DecimalField(max_digits=5, decimal_places=2)
        )
    )
    return trattative_con_margine.order_by('-data_ultimo_aggiornamento')


@login_required 
def dashboard(request):
    # ... (Il resto della vista dashboard è invariato) ...
    selected_year = request.GET.get('anno'); selected_month = request.GET.get('mese')
    vendite_qs = Vendita.objects.all(); trattative_qs = Trattativa.objects.all(); stats_qs = StatMensile.objects.all()
    filter_title = "Totale Complessivo"; is_filtered = False
    if selected_year:
        vendite_qs = vendite_qs.filter(data_vendita__year=selected_year); trattative_qs = trattative_qs.filter(data_creazione__year=selected_year); stats_qs = stats_qs.filter(anno=selected_year)
        filter_title = f"Anno {selected_year}"; is_filtered = True
    if selected_month:
        vendite_qs = vendite_qs.filter(data_vendita__month=selected_month); trattative_qs = trattative_qs.filter(data_creazione__month=selected_month); stats_qs = stats_qs.filter(mese=selected_month)
        filter_title = f"{selected_month}/{selected_year}"; is_filtered = True

    trattative_vinte_nel_periodo = Trattativa.objects.filter(vendita_collegata__in=vendite_qs)
    
    aggregati_vendite = vendite_qs.aggregate(
        tot_venduto_prodotti=Coalesce(Sum('prezzo_vendita'), Decimal(0)),
        tot_costo_prodotti=Coalesce(Sum('costo_acquisto'), Decimal(0)),
        numero_vendite=Count('id'),
        numero_finanziamenti=Count('id', filter=Q(flag_finanziamento=True)),
        numero_resi=Coalesce(Count('id', filter=Q(flag_reso=True)), 0)
    )
    
    aggregati_servizi = Attivita.objects.filter(
        trattativa__in=trattative_vinte_nel_periodo
    ).aggregate(
        tot_venduto_servizi=Coalesce(Sum('prezzo_vendita_attivita'), Decimal(0)),
        costo_personale_periodo=Coalesce(get_costo_attivita_query(), Decimal(0.0))
    )
    
    vendite_totali = aggregati_vendite['tot_venduto_prodotti'] + aggregati_servizi['tot_venduto_servizi']
    costi_totali_cogs = aggregati_vendite['tot_costo_prodotti'] 
    margine_lordo_euro = vendite_totali - costi_totali_cogs
    
    margine_totale_percent = Decimal(0)
    if vendite_totali > 0: margine_totale_percent = (margine_lordo_euro / vendite_totali) * 100
    numero_vendite = aggregati_vendite['numero_vendite']
    scontrino_medio = Decimal(0)
    if numero_vendite > 0: scontrino_medio = vendite_totali / numero_vendite
    perc_finanziamenti = Decimal(0)
    if numero_vendite > 0: perc_finanziamenti = (aggregati_vendite['numero_finanziamenti'] / Decimal(numero_vendite)) * 100
    
    perc_incidenza_servizi = Decimal(0)
    if vendite_totali > 0:
        perc_incidenza_servizi = (aggregati_servizi['tot_venduto_servizi'] / vendite_totali) * 100

    costi_manuali = stats_qs.aggregate(
        tot_fissi=Coalesce(Sum('costi_operativi_fissi'), Decimal(0)),
        tot_marketing=Coalesce(Sum('costo_marketing_mese'), Decimal(0)),
        tot_finan_respinti=Coalesce(Sum('finanziamenti_non_approvati'), 0)
    )
    costi_fissi_periodo = costi_manuali['tot_fissi']
    costo_personale_periodo = aggregati_servizi['costo_personale_periodo']
    
    costi_operativi_totali = costi_fissi_periodo + costo_personale_periodo + costi_manuali['tot_marketing']
    utile_operativo_ebit = margine_lordo_euro - costi_operativi_totali

    aggregati_trattative = trattative_qs.aggregate(
        lead_generati=Count('id'),
        preventivi_persi=Count('id', filter=Q(stato=Trattativa.STATO_PERSO))
    )
    perc_preventivi_persi = Decimal(0)
    if aggregati_trattative['lead_generati'] > 0:
        perc_preventivi_persi = (aggregati_trattative['preventivi_persi'] / Decimal(aggregati_trattative['lead_generati'])) * 100
    costo_per_lead = Decimal(0)
    if aggregati_trattative['lead_generati'] > 0:
        costo_per_lead = costi_manuali['tot_marketing'] / Decimal(aggregati_trattative['lead_generati'])
    tempo_evasione_avg = trattative_vinte_nel_periodo.annotate(
        tempo=ExpressionWrapper(F('vendita_collegata__data_vendita') - F('data_creazione'), output_field=models.DurationField())
    ).aggregate(
        tempo_medio=Avg('tempo')
    )['tempo_medio']
    tempo_medio_evasione_giorni = tempo_evasione_avg.days if tempo_evasione_avg else 0
    
    costo_montaggio = Attivita.objects.filter(
        trattativa__in=trattative_vinte_nel_periodo, 
        categoria__nome__icontains='Montaggio'
    ).aggregate(
        costo_totale=Coalesce(get_costo_attivita_query(), Decimal(0.0))
    )['costo_totale']
    costo_medio_montaggio = Decimal(0)
    if numero_vendite > 0:
        costo_medio_montaggio = costo_montaggio / Decimal(numero_vendite)

    budget_lookup = {}
    if selected_year and selected_month:
        budget_qs = Budget.objects.filter(anno=selected_year, mese=selected_month)
        for b in budget_qs: budget_lookup[b.categoria_id] = {'vendite': b.obiettivo_vendite_euro, 'margine_perc': b.obiettivo_margine_percentuale}
    
    categorie_summary_qs = CategoriaMerceologica.objects.annotate(
        tot_venduto=Coalesce(Sum('vendita__prezzo_vendita', filter=Q(vendita__in=vendite_qs)), Decimal(0)),
        tot_costo=Coalesce(Sum('vendita__costo_acquisto', filter=Q(vendita__in=vendite_qs)), Decimal(0))
    ).annotate(margine_euro=F('tot_venduto') - F('tot_costo')).annotate(
        margine_perc=Case(
            When(tot_venduto=0, then=Value(Decimal(0.0))),
            default=ExpressionWrapper((F('margine_euro') * Decimal('100.0') / F('tot_venduto')), output_field=DecimalField(max_digits=5, decimal_places=2)),
            output_field=DecimalField(max_digits=5, decimal_places=2)
        )
    ).filter(tot_venduto__gt=0).order_by('-tot_venduto')
    
    categorie_summary = []
    alerts = [] 
    for cat in categorie_summary_qs:
        budget_cat = budget_lookup.get(cat.id)
        scostamento_vendite, scostamento_margine = None, None
        if budget_cat:
            scostamento_vendite = cat.tot_venduto - budget_cat['vendite']
            scostamento_margine = (cat.margine_perc or Decimal(0)) - budget_cat['margine_perc']
            if scostamento_margine < -5: alerts.append({'level': 'danger', 'message': f"Margine Categoria '{cat.nome}' in forte calo: {cat.margine_perc:.2f}% (Budget: {budget_cat['margine_perc']}%)"})
            elif scostamento_margine < 0: alerts.append({'level': 'warning', 'message': f"Margine Categoria '{cat.nome}' sotto budget: {cat.margine_perc:.2f}% (Budget: {budget_cat['margine_perc']}%)"})
        categorie_summary.append({'nome': cat.nome, 'tot_venduto': cat.tot_venduto, 'margine_euro': cat.margine_euro, 'margine_perc': cat.margine_perc, 'budget_vendite': budget_cat['vendite'] if budget_cat else None, 'budget_margine_perc': budget_cat['margine_perc'] if budget_cat else None, 'scostamento_vendite': scostamento_vendite, 'scostamento_margine': scostamento_margine})
    
    if tempo_medio_evasione_giorni > ALERT_SOGLIA_GIORNI_EVASIONE: alerts.append({'level': 'warning', 'message': f"Tempo medio evasione ordini OLTRE SOGLIA: {tempo_medio_evasione_giorni} giorni"})
    if aggregati_vendite['numero_resi'] > ALERT_SOGLIA_RESI_DIFETTI: alerts.append({'level': 'danger', 'message': f"Aumento Resi/Difetti: {aggregati_vendite['numero_resi']} casi rilevati"})
    if costi_manuali['tot_finan_respinti'] > ALERT_SOGLIA_FINAN_NON_APPROV: alerts.append({'level': 'info', 'message': f"Finanziamenti non approvati in aumento: {costi_manuali['tot_finan_respinti']} casi"})

    chart_labels, chart_data_vendite, chart_data_margine, chart_data_ebit = [], [], [], []
    if not is_filtered:
        today = timezone.now().date(); start_date = today - datetime.timedelta(days=365)
        stats_costs_q = StatMensile.objects.filter(anno=OuterRef('month__year'), mese=OuterRef('month__month'))
        stats_costs_fissi = stats_costs_q.values('costi_operativi_fissi'); stats_costs_mktg = stats_costs_q.values('costo_marketing_mese')
        
        trend_prodotti = Vendita.objects.filter(data_vendita__gte=start_date).annotate(month=TruncMonth('data_vendita')).values('month').annotate(
            tot_venduto_prod=Coalesce(Sum('prezzo_vendita'), Decimal(0)), 
            tot_costo_prod=Coalesce(Sum('costo_acquisto'), Decimal(0)), 
            costi_fissi=Coalesce(Subquery(stats_costs_fissi[:1]), Decimal(0)), 
            costi_mktg=Coalesce(Subquery(stats_costs_mktg[:1]), Decimal(0))
        ).order_by('month')
        
        trend_servizi = Attivita.objects.filter(
            trattativa__vendita_collegata__data_vendita__gte=start_date
        ).annotate(
            month=TruncMonth('trattativa__vendita_collegata__data_vendita')
        ).values('month').annotate(
            tot_venduto_serv=Coalesce(Sum('prezzo_vendita_attivita'), Decimal(0)),
            costo_personale=Coalesce(get_costo_attivita_query(), Decimal(0.0))
        ).order_by('month')
        
        prodotti_dict = {d['month']: d for d in trend_prodotti}
        servizi_dict = {s['month']: s for s in trend_servizi}
        all_months = sorted(list(set(prodotti_dict.keys()) | set(servizi_dict.keys())))

        for month in all_months:
            data_prod = prodotti_dict.get(month, {})
            data_serv = servizi_dict.get(month, {})
            
            month_label = month.strftime('%m/%Y')
            vendite_prodotti = data_prod.get('tot_venduto_prod', Decimal(0))
            vendite_servizi = data_serv.get('tot_venduto_serv', Decimal(0))
            vendite_mese = vendite_prodotti + vendite_servizi
            
            costo_prodotti = data_prod.get('tot_costo_prod', Decimal(0))
            margine_mese = vendite_mese - costo_prodotti
            
            costi_fissi_mese = data_prod.get('costi_fissi', Decimal(0))
            costo_mktg_mese = data_prod.get('costi_mktg', Decimal(0))
            costo_personale_mese = data_serv.get('costo_personale', Decimal(0))
            
            ebit_mese = margine_mese - costi_fissi_mese - costo_personale_mese - costo_mktg_mese
            
            chart_labels.append(month_label)
            chart_data_vendite.append(float(vendite_mese))
            chart_data_margine.append(float(margine_mese))
            chart_data_ebit.append(float(ebit_mese))

    stati_attivi = [Trattativa.STATO_LEAD, Trattativa.STATO_APPUNTAMENTO, Trattativa.STATO_PROGETTAZIONE, Trattativa.STATO_PREVENTIVO, Trattativa.STATO_CONSEGNA, Trattativa.STATO_MONTAGGIO] 
    pipeline_attiva_qs = get_trattative_annotate().filter(stato__in=stati_attivi) 
    
    pipeline_aggregati = pipeline_attiva_qs.aggregate(
        valore_totale=Coalesce(Sum('valore_totale_stimato'), Decimal(0)), 
        numero_trattative=Count('id'),
        costo_personale_loggato=Coalesce(Sum('costo_personale_totale'), Decimal(0))
    )
    pipeline_valore = pipeline_aggregati['valore_totale']
    pipeline_numero = pipeline_aggregati['numero_trattative']
    pipeline_costo_loggato = pipeline_aggregati['costo_personale_loggato'] 
            
    available_years = Vendita.objects.dates('data_vendita', 'year', order='DESC'); available_months = range(1, 13)
    context = {
        'vendite_totali': vendite_totali, 'margine_totale_euro': margine_lordo_euro, 'margine_totale_percent': margine_totale_percent,
        'utile_operativo_ebit': utile_operativo_ebit, 'scontrino_medio': scontrino_medio, 'perc_incidenza_servizi': perc_incidenza_servizi,
        'perc_finanziamenti': perc_finanziamenti, 'numero_resi': aggregati_vendite['numero_resi'], 'numero_vendite': numero_vendite,
        'pipeline_valore': pipeline_valore, 'pipeline_numero': pipeline_numero, 'pipeline_costo_loggato': pipeline_costo_loggato,
        'costo_personale_periodo': costo_personale_periodo, 'kpi_perc_preventivi_persi': perc_preventivi_persi,
        'kpi_costo_per_lead': costo_per_lead, 'kpi_costo_medio_montaggio': costo_medio_montaggio,
        'kpi_tempo_medio_evasione_giorni': tempo_medio_evasione_giorni, 'kpi_finanziamenti_non_approvati': costi_manuali['tot_finan_respinti'],
        'categorie_summary': categorie_summary, 'filter_title': filter_title, 'is_filtered': is_filtered,
        'available_years': available_years, 'available_months': available_months, 'selected_year': int(selected_year) if selected_year else None,
        'selected_month': int(selected_month) if selected_month else None, 'active_page': 'dashboard', 'alerts': alerts,
        'chart_labels': json.dumps(chart_labels), 'chart_data_vendite': json.dumps(chart_data_vendite),
        'chart_data_margine': json.dumps(chart_data_margine), 'chart_data_ebit': json.dumps(chart_data_ebit),
    }
    return render(request, 'gestione/dashboard.html', context)


@login_required
def report_venditori(request):
    # ... (Il resto della vista report_venditori è invariato) ...
    selected_year = request.GET.get('anno'); selected_month = request.GET.get('mese')
    vendite_qs = Vendita.objects.all(); filter_title = "Totale Complessivo" 
    if selected_year: vendite_qs = vendite_qs.filter(data_vendita__year=selected_year); filter_title = f"Anno {selected_year}"
    if selected_month: vendite_qs = vendite_qs.filter(data_vendita__month=selected_month); filter_title = f"{selected_month}/{selected_year}"
    sales_report = User.objects.filter(vendita__in=vendite_qs).distinct().annotate(
        tot_venduto=Coalesce(Sum('vendita__prezzo_vendita', filter=Q(vendita__in=vendite_qs)), Decimal(0)),
        tot_costo=Coalesce(Sum('vendita__costo_acquisto', filter=Q(vendita__in=vendite_qs)), Decimal(0)),
        num_vendite=Count('vendita__id', filter=Q(vendita__in=vendite_qs))
    ).annotate(margine_euro=F('tot_venduto') - F('tot_costo')).annotate(
        margine_perc=Case(When(tot_venduto=0, then=Value(Decimal(0.0))), default=ExpressionWrapper((F('margine_euro') * Decimal('100.0') / F('tot_venduto')), output_field=DecimalField(max_digits=5, decimal_places=2))),
        scontrino_medio=Case(When(num_vendite=0, then=Value(Decimal(0.0))), default=ExpressionWrapper(F('tot_venduto') / F('num_vendite'), output_field=DecimalField(max_digits=10, decimal_places=2)))
    ).filter(num_vendite__gt=0).order_by('-tot_venduto') 
    available_years = Vendita.objects.dates('data_vendita', 'year', order='DESC'); available_months = range(1, 13)
    context = {
        'sales_report': sales_report, 'filter_title': filter_title, 'available_years': available_years, 'available_months': available_months,
        'selected_year': int(selected_year) if selected_year else None, 'selected_month': int(selected_month) if selected_month else None,
        'active_page': 'venditori', 
    }
    return render(request, 'gestione/report_venditori.html', context)


# --- VISTE KANBAN ---
@login_required
def kanban_board(request):
    # ... (Il resto della vista kanban_board è invariato) ...
    stati_kanban = [
        (Trattativa.STATO_LEAD, '1. Lead/Contatto'), (Trattativa.STATO_APPUNTAMENTO, '2. Appuntamento Fissato'),
        (Trattativa.STATO_PROGETTAZIONE, '3. Progettazione'), (Trattativa.STATO_PREVENTIVO, '4. Preventivo Inviato'),
        (Trattativa.STATO_CONSEGNA, '5. In Consegna'), (Trattativa.STATO_MONTAGGIO, '6. In Montaggio'),
        (Trattativa.STATO_VINTO, '7. Chiuso Vinto'), (Trattativa.STATO_PERSO, '8. Chiuso Perso'),
    ]
    trattative_annotate = get_trattative_annotate().filter(
        stato__in=[s[0] for s in stati_kanban]
    )
    colonne = []
    for stato_key, stato_display in stati_kanban:
        colonne.append({
            'stato_key': stato_key, 
            'stato_display': stato_display,
            'trattative': [t for t in trattative_annotate if t.stato == stato_key]
        })
    context = {'active_page': 'kanban', 'colonne': colonne,}
    return render(request, 'gestione/kanban_board.html', context)


@csrf_exempt
@login_required
@require_POST
def move_trattativa(request):
    # ... (Il resto della vista move_trattativa è invariato) ...
    try:
        trattativa_id = request.POST.get('id'); nuovo_stato = request.POST.get('stato')
        trattativa = get_object_or_404(Trattativa, pk=trattativa_id)
        if trattativa.stato == Trattativa.STATO_VINTO or trattativa.stato == Trattativa.STATO_PERSO:
            return HttpResponse(status=403, content="Trattativa già chiusa.")
        trattativa.stato = nuovo_stato; trattativa.save()
        headers = {'HX-Trigger': json.dumps({'trattativaMossa': {'trattativaId': trattativa.id,'nuovoStato': nuovo_stato}})}
        return HttpResponse(status=204, headers=headers)
    except Exception as e:
        return HttpResponse(status=400, content=str(e))

@login_required
def chiudi_trattativa_modal(request, trattativa_id):
    # ... (Il resto della vista chiudi_trattativa_modal è invariato) ...
    trattativa = get_object_or_404(Trattativa, pk=trattativa_id)
    if request.method == 'POST':
        form = TrattativaVintaForm(request.POST)
        if form.is_valid():
            vendita = form.save(commit=False); vendita.venditore = trattativa.commerciale or request.user; vendita.cliente = trattativa.cliente_nome; vendita.save()
            trattativa.stato = Trattativa.STATO_VINTO; trattativa.vendita_collegata = vendita; trattativa.valore_stimato = vendita.prezzo_vendita; trattativa.costo_materiali_stimato = vendita.costo_acquisto; trattativa.save()
            messages.success(request, f"Trattativa '{trattativa.titolo}' chiusa con successo!")
            headers = {'HX-Trigger': json.dumps({'trattativaChiusa': {'trattativaId': trattativa.id}})}
            return HttpResponse(status=204, headers=headers)
        else:
            messages.error(request, "Form non valido.")
    else: 
        form = TrattativaVintaForm(initial={
            'descrizione': trattativa.titolo, 'cliente': trattativa.cliente_nome, 
            'prezzo_vendita': trattativa.valore_stimato, 'costo_acquisto': trattativa.costo_materiali_stimato,
            'data_vendita': datetime.date.today(),
        })
    return render(request, 'gestione/partials/modal_chiudi_vinto.html', {'form': form, 'trattativa': trattativa})


# --- VISTA report_attivita (CORRETTA) ---
@login_required
def report_attivita(request):
    # --- CORREZIONE QUI ---
    # Sostituiamo "costo_per_commerciale" con "costo_per_ruolo"
    costo_per_ruolo = RuoloCosto.objects.filter(attivita__isnull=False).distinct().annotate(
        ore_totali=Coalesce(Sum('attivita__tempo_dedicato_ore'), Decimal(0.0))
    ).annotate(
        # Calcoliamo il costo totale moltiplicando le ore per il costo orario del ruolo
        costo_totale=F('ore_totali') * F('costo_orario')
    ).order_by('-costo_totale')
    # --- FINE CORREZIONE ---

    stati_attivi = [Trattativa.STATO_LEAD, Trattativa.STATO_APPUNTAMENTO, Trattativa.STATO_PROGETTAZIONE, Trattativa.STATO_PREVENTIVO, Trattativa.STATO_CONSEGNA, Trattativa.STATO_MONTAGGIO]
    
    # Questa query è corretta perché get_costo_personale_query usa 'attivita__ruolo__costo_orario'
    costo_per_trattativa_attiva = Trattativa.objects.filter(stato__in=stati_attivi, attivita__isnull=False).distinct().annotate(
        ore_totali=Coalesce(Sum('attivita__tempo_dedicato_ore'), Decimal(0.0)),
        costo_totale=Coalesce(get_costo_personale_query(), Decimal(0.0)) # Questa funzione è corretta
    ).select_related('commerciale').order_by('-costo_totale')
    
    costo_trattative_perse = Attivita.objects.filter(trattativa__stato=Trattativa.STATO_PERSO).aggregate(costo_totale_perso=Coalesce(get_costo_attivita_query(), Decimal(0.0)))['costo_totale_perso']
    costo_trattative_vinte = Attivita.objects.filter(trattativa__stato=Trattativa.STATO_VINTO).aggregate(costo_totale_vinte=Coalesce(get_costo_attivita_query(), Decimal(0.0)))['costo_totale_vinte']
    
    context = {
        'active_page': 'report_attivita', 
        'costo_per_ruolo': costo_per_ruolo, # <-- Nome variabile cambiato
        'costo_per_trattativa_attiva': costo_per_trattativa_attiva,
        'costo_trattative_perse': costo_trattative_perse, 
        'costo_trattative_vinte': costo_trattative_vinte,
    }
    return render(request, 'gestione/report_attivita.html', context)


# --- VISTE PAGINA DETTAGLIO ---
@login_required
def trattativa_dettaglio(request, trattativa_id):
    trattativa = get_object_or_404(Trattativa, pk=trattativa_id)
    if request.method == 'POST':
        form_dati = TrattativaDettaglioForm(request.POST, instance=trattativa)
        if form_dati.is_valid():
            form_dati.save()
            messages.success(request, "Dati trattativa aggiornati.")
            return redirect('trattativa_dettaglio', trattativa_id=trattativa.id)
        else:
            messages.error(request, "Errore nel salvataggio dei dati.")
    else:
        form_dati = TrattativaDettaglioForm(instance=trattativa)
    
    # Query corretta (usa 'ruolo')
    attivita = trattativa.attivita.all().select_related('categoria', 'ruolo').order_by('-data_attivita')
    
    messaggi_chat = trattativa.messaggi_chat.all().select_related('utente').order_by('timestamp')
    
    # Calcolo corretto (usa 'get_costo_personale_trattativa')
    costo_personale_trattativa = get_costo_personale_trattativa(trattativa_id)
    ricavo_servizi_trattativa = attivita.aggregate(total=Coalesce(Sum('prezzo_vendita_attivita'), Decimal(0.0)))['total']
    
    valore_totale = trattativa.valore_stimato + ricavo_servizi_trattativa
    costo_totale = trattativa.costo_materiali_stimato + costo_personale_trattativa
    
    margine_stimato_euro = valore_totale - costo_totale
    margine_stimato_perc = Decimal(0)
    if valore_totale > 0:
        margine_stimato_perc = (margine_stimato_euro / valore_totale) * 100
    
    attivita_form = AttivitaForm()
    chat_form = MessaggioChatForm()
    context = {
        'active_page': 'kanban', 'trattativa': trattativa, 'form_dati': form_dati,
        'attivita_lista': attivita, 'messaggi_chat': messaggi_chat,
        'attivita_form': attivita_form, 'chat_form': chat_form,
        'valore_totale_stimato': valore_totale,
        'costo_personale_trattativa': costo_personale_trattativa,
        'ricavo_servizi_trattativa': ricavo_servizi_trattativa,
        'margine_stimato_euro': margine_stimato_euro, 
        'margine_stimato_perc': margine_stimato_perc,
    }
    return render(request, 'gestione/trattativa_dettaglio.html', context)


# --- VISTA add_attivita (SEMPLIFICATA) ---
@login_required
def add_attivita(request, trattativa_id):
    trattativa = get_object_or_404(Trattativa, pk=trattativa_id)
    
    if request.method == 'POST':
        form = AttivitaForm(request.POST)
        if form.is_valid():
            form.save(commit=False)
            form.instance.trattativa = trattativa
            form.save()
            messages.success(request, "Attività loggata con successo!")
            headers = {'HX-Refresh': 'true'}
            return HttpResponse(status=204, headers=headers)
        else:
            messages.error(request, "Errore nel form attività.")
            return render(request, 'gestione/partials/_modal_add_attivita.html', {
                'trattativa': trattativa, 
                'attivita_form_errors': form.errors, 
                'attivita_form': form,         
                'is_editing': False
            }, status=400)
    
    # --- LOGICA GET ---
    else: 
        form = AttivitaForm(initial={
            'data_attivita': datetime.date.today(),
        })
        
        return render(request, 'gestione/partials/_modal_add_attivita.html', {
            'trattativa': trattativa, 
            'attivita_form': form,
            'is_editing': False
        })

# --- VISTA: EDIT ATTIVITA (SEMPLIFICATA) ---
@login_required
def edit_attivita(request, attivita_id):
    attivita = get_object_or_404(Attivita, pk=attivita_id)
    trattativa = attivita.trattativa
    
    if request.method == 'POST':
        form = AttivitaForm(request.POST, instance=attivita)
        if form.is_valid():
            form.save()
            messages.success(request, "Attività aggiornata con successo!")
            return HttpResponse(status=204, headers={'HX-Refresh': 'true'})
        else:
            messages.error(request, "Errore nel form attività.")
            return render(request, 'gestione/partials/_modal_add_attivita.html', {
                'trattativa': trattativa,
                'attivita_form': form,
                'attivita_form_errors': form.errors,
                'is_editing': True
            }, status=400)
    
    # --- LOGICA GET ---
    else:
        form = AttivitaForm(instance=attivita)
        return render(request, 'gestione/partials/_modal_add_attivita.html', {
            'trattativa': trattativa,
            'attivita_form': form,
            'is_editing': True
        })

# --- VISTA: DELETE ATTIVITA ---
@login_required
@require_POST
def delete_attivita(request, attivita_id):
    try:
        attivita = get_object_or_404(Attivita, pk=attivita_id)
        attivita.delete()
        messages.success(request, "Attività eliminata con successo.")
        # Ricarica l'intera pagina per aggiornare i totali
        return HttpResponse(status=204, headers={'HX-Refresh': 'true'})
    except Exception as e:
        messages.error(request, f"Errore durante l'eliminazione: {e}")
        return HttpResponse(status=400, headers={'HX-Refresh': 'true'})


# --- NUOVA VISTA PER CALCOLO COSTI LIVE (HTMX) ---
@login_required
def calcola_costi_attivita(request):
    
    # 1. Recupera i dati inviati da HTMX
    ruolo_id = request.GET.get('ruolo')
    ore_str = request.GET.get('tempo_dedicato_ore', '0').replace(',', '.')
    prezzo_str = request.GET.get('prezzo_vendita_attivita', '0').replace(',', '.')

    costo_orario = Decimal(0.0)
    costo_totale = Decimal(0.0)
    ore = Decimal(0.0)
    prezzo_vendita = Decimal(0.0)
    
    show_alert = False
    alert_message = ""

    # 2. Cerca il costo del ruolo
    if ruolo_id:
        try:
            ruolo = RuoloCosto.objects.get(pk=ruolo_id)
            if ruolo.costo_orario is not None:
                costo_orario = ruolo.costo_orario
        except RuoloCosto.DoesNotExist:
            pass # costo_orario rimane 0

    # 3. Fai i calcoli
    try:
        ore = Decimal(ore_str) if ore_str else Decimal(0.0)
    except Exception:
        ore = Decimal(0.0)
        
    try:
        prezzo_vendita = Decimal(prezzo_str) if prezzo_str else Decimal(0.0)
    except Exception:
        prezzo_vendita = Decimal(0.0)

    costo_totale = ore * costo_orario

    # 4. Controlla il margine (logica degli alert)
    if costo_totale > 0:
        impostazioni = ImpostazioniGenerali.load()
        soglia_margine = impostazioni.soglia_alert_margine_servizio / Decimal(100.0)
        
        prezzo_minimo = costo_totale * (1 + soglia_margine)

        if prezzo_vendita <= 0:
            show_alert = True
            alert_message = f"(Stai offrendo un servizio a €0.00 che costa € {costo_totale:.2f})"
        elif prezzo_vendita < prezzo_minimo:
            show_alert = True
            alert_message = f"(Prezzo min. suggerito: € {prezzo_minimo:.2f})"
            
    context = {
        'costo_orario': costo_orario,
        'costo_totale': costo_totale,
        'show_alert': show_alert,
        'alert_message': alert_message,
    }
    
    # 5. Restituisci il partial template
    return render(request, 'gestione/partials/_partial_calcolo_costi.html', context)


@login_required
@require_POST
def add_messaggio(request, trattativa_id):
    trattativa = get_object_or_404(Trattativa, pk=trattativa_id)
    form = MessaggioChatForm(request.POST)
    if form.is_valid():
        messaggio = form.save(commit=False)
        messaggio.trattativa = trattativa
        messaggio.utente = request.user
        messaggio.save()
        return render(request, 'gestione/partials/_partial_singolo_messaggio.html', {'msg': messaggio})
    messages.error(request, "Errore: il messaggio non può essere vuoto.")
    return HttpResponse(status=400)


@login_required
def statistiche_mensili(request):
    anno_corrente = datetime.date.today().year
    mese_corrente = datetime.date.today().month
    anno = request.GET.get('anno', anno_corrente)
    mese = request.GET.get('mese', mese_corrente)
    try:
        anno = int(anno)
        mese = int(mese)
    except (ValueError, TypeError):
        anno = anno_corrente
        mese = mese_corrente
    stat_mensile, created = StatMensile.objects.get_or_create(anno=anno, mese=mese, defaults={})
    if request.method == 'POST':
        form = StatMensileForm(request.POST, instance=stat_mensile)
        if form.is_valid():
            form.save()
            messages.success(request, f"Dati manuali per {mese}/{anno} salvati con successo!")
            return redirect(f"{request.path}?anno={anno}&mese={mese}")
        else:
            messages.error(request, "Errore nella compilazione del form.")
    else:
        form = StatMensileForm(instance=stat_mensile)
    anni_disponibili = range(anno_corrente - 3, anno_corrente + 2)
    context = {
        'form': form, 'anno_selezionato': anno, 'mese_selezionato': mese,
        'anni_disponibili': anni_disponibili, 'active_page': 'statistiche_mensili',
    }
    return render(request, 'gestione/statistiche_mensili.html', context)


@login_required
def nuova_trattativa(request):
    if request.method == 'POST':
        form = TrattativaForm(request.POST)
        if form.is_valid():
            trattativa = form.save(commit=False)
            trattativa.stato = Trattativa.STATO_LEAD
            if not trattativa.commerciale:
                trattativa.commerciale = request.user
            trattativa.save()
            messages.success(request, f"Trattativa '{trattativa.titolo}' creata con successo!")
            return redirect('kanban_board')
        else:
            messages.error(request, "Errore nella compilazione del form.")
    else:
        form = TrattativaForm(initial={'commerciale': request.user})
    context = {'form': form, 'active_page': 'kanban',}
    return render(request, 'gestione/nuova_trattativa.html', context)


# --- NUOVE VISTE PER PAGINA LISTA (PUNTO 17) ---

@login_required
def trattativa_lista(request):
    """
    Mostra una vista tabellare di tutte le trattative
    con i costi e i margini calcolati.
    """
    trattative_list = get_trattative_annotate().order_by('-data_ultimo_aggiornamento')
    
    context = {
        'active_page': 'lista', # Per la navbar
        'trattative': trattative_list
    }
    return render(request, 'gestione/trattativa_lista.html', context)


@login_required
def esporta_trattative_excel(request):
    """
    Genera e scarica un file Excel con il report
    completo di tutte le trattative.
    """
    trattative = get_trattative_annotate().order_by('stato', '-data_ultimo_aggiornamento')
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report Trattative"

    # Intestazioni
    headers = [
        "ID", "Titolo", "Stato", "Cliente", "Commerciale", 
        "Data Creazione", "Valore Prodotto (€)", "Ricavo Servizi (€)", "Valore Totale (€)",
        "Costo Materiali (€)", "Costo Personale (€)", "Costo Totale (€)",
        "Margine Stimato (€)", "Margine Stimato (%)"
    ]
    ws.append(headers)

    # Applica stile alle intestazioni
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Popola i dati
    for t in trattative:
        commerciale_username = t.commerciale.username if t.commerciale else "N/A"
        
        ws.append([
            t.id, t.titolo, t.get_stato_display(), t.cliente_nome, commerciale_username,
            t.data_creazione.replace(tzinfo=None), # Rimuovi timezone per Excel
            float(t.valore_stimato), # Valore Prodotto
            float(t.ricavo_servizi_totale), # Ricavo Servizi
            float(t.valore_totale_stimato), # Valore Totale
            float(t.costo_materiali_stimato), # Costo Materiali
            float(t.costo_personale_totale), # Costo Personale
            float(t.costo_totale_stimato), # Costo Totale
            float(t.margine_stimato_euro), # Margine €
            float(t.margine_stimato_perc) / 100.0 # Margine %
        ])

    # Aggiusta la larghezza delle colonne e formatta i numeri
    for i, col in enumerate(headers, 1):
        col_letter = get_column_letter(i)
        if "€" in col:
            ws.column_dimensions[col_letter].width = 18
            for cell in ws[col_letter][1:]: # Salta l'intestazione
                cell.number_format = '€ #,##0.00'
        elif "%" in col:
            ws.column_dimensions[col_letter].width = 15
            for cell in ws[col_letter][1:]:
                cell.number_format = '0.00%'
        else:
            ws.column_dimensions[col_letter].width = 20 # Larghezza standard

    # Crea la risposta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="report_trattative_{timezone.now().strftime("%Y-%m-%d")}.xlsx"'
    
    wb.save(response)
    return response